import logging
from decimal import Decimal

from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.common.audit import record as audit
from apps.common.views import SchoolScopedViewSet
from apps.schools.models import School

from .models import (
    DEFAULT_VOTE_HEADS,
    FeeStructure,
    Invoice,
    Payment,
    StkPushRequest,
)
from .serializers import (
    FeeStructureSerializer,
    InvoiceSerializer,
    PaymentSerializer,
    StkPushSerializer,
)
from .services import daraja
from .services.reconcile import record_transaction

logger = logging.getLogger(__name__)


class FeeStructureViewSet(SchoolScopedViewSet):
    """What a class is charged per term. Staff read it; the office sets it —
    the amount here becomes every learner's invoice."""

    queryset = FeeStructure.objects.all()
    serializer_class = FeeStructureSerializer
    filterset_fields = ["grade", "term", "year"]

    def _require_office(self):
        user = self.request.user
        if not (user.is_superuser or user.role == "ADMIN"):
            raise PermissionDenied("Only the school office sets fee structures.")

    def perform_create(self, serializer):
        self._require_office()
        super().perform_create(serializer)

    def perform_update(self, serializer):
        self._require_office()
        serializer.save()

    def perform_destroy(self, instance):
        self._require_office()
        instance.delete()


class InvoiceViewSet(SchoolScopedViewSet):
    queryset = (
        Invoice.objects.select_related("learner", "fee_structure")
        .prefetch_related("payments")
        .all()
    )
    serializer_class = InvoiceSerializer
    # The office slices the register by class and by term:
    #   ?learner__grade=5&learner__stream=North&fee_structure__term=2
    filterset_fields = [
        "learner", "status", "fee_structure", "learner__grade",
        "learner__stream", "fee_structure__term", "fee_structure__year",
    ]
    ordering_fields = ["learner__admission_number", "amount_due", "amount_paid", "status"]
    search_fields = ["learner__first_name", "learner__last_name",
                     "learner__admission_number"]

    def get_queryset(self):
        return super().get_queryset().order_by(
            "learner__grade", "learner__stream", "learner__admission_number"
        )


class PaymentViewSet(SchoolScopedViewSet):
    """The instalment ledger. Recording or reversing a payment re-totals its
    invoice, so paid/balance/status always follow the rows beneath them."""

    queryset = Payment.objects.select_related(
        "invoice__learner", "received_by"
    ).all()
    serializer_class = PaymentSerializer
    filterset_fields = ["invoice", "method", "paid_on"]

    def _require_office(self):
        user = self.request.user
        if not (user.is_superuser or user.role == "ADMIN"):
            raise PermissionDenied("Fee payments are recorded by the school office.")

    def _retotal(self, invoice):
        total = invoice.payments.aggregate(t=Sum("amount"))["t"] or Decimal("0")
        invoice.amount_paid = total
        invoice.refresh_status()

    def perform_create(self, serializer):
        self._require_office()
        payment = serializer.save(
            school=self.request.user.school, received_by=self.request.user
        )
        self._retotal(payment.invoice)
        audit(
            actor=self.request.user,
            school=payment.school,
            action="PAYMENT_RECORDED",
            target=payment,
            label=f"{payment.invoice.learner.full_name} — KES {payment.amount}",
            detail={"method": payment.method, "reference": payment.reference},
        )

    def perform_update(self, serializer):
        self._require_office()
        payment = serializer.save()
        self._retotal(payment.invoice)

    def perform_destroy(self, instance):
        self._require_office()
        invoice = instance.invoice
        instance.delete()
        self._retotal(invoice)


class FeeStructureGridView(APIView):
    """GET/POST /api/payments/fee-structures/grid/ — the fee note as a school
    prints it: grades down the side, vote heads across the top.

    GET  ?term=&year= → the columns and one row per grade with its amounts.
    POST {term, year, columns, rows:[{grade, breakdown}]} → upsert the lot.
    A grade whose row totals zero is dropped, so clearing a row removes it.
    """

    def _require_office(self, user):
        if not (user.is_superuser or user.role == "ADMIN"):
            raise PermissionDenied("Only the school office sets fee structures.")

    def get(self, request):
        from apps.schools.moe import ALL_GRADES, GRADE_LABELS

        school = request.user.school
        try:
            term = int(request.query_params.get("term"))
            year = int(request.query_params.get("year"))
        except (TypeError, ValueError):
            raise ValidationError({"detail": "term and year are required."})

        existing = {
            fs.grade: fs
            for fs in FeeStructure.objects.filter(school=school, term=term, year=year)
        }
        # Whatever columns the school has actually used, then the defaults.
        used = []
        for fs in existing.values():
            for head in (fs.breakdown or {}):
                if head not in used:
                    used.append(head)
        columns = used + [h for h in DEFAULT_VOTE_HEADS if h not in used]

        return Response({
            "term": term,
            "year": year,
            "columns": columns,
            "rows": [
                {
                    "grade": g,
                    "label": GRADE_LABELS[g],
                    "breakdown": (existing[g].breakdown or {}) if g in existing else {},
                    "total": str(existing[g].amount) if g in existing else "0",
                }
                for g in ALL_GRADES
            ],
        })

    def post(self, request):
        self._require_office(request.user)
        school = request.user.school
        try:
            term = int(request.data.get("term"))
            year = int(request.data.get("year"))
        except (TypeError, ValueError):
            raise ValidationError({"detail": "term and year are required."})

        saved = cleared = 0
        for row in request.data.get("rows", []):
            try:
                grade = int(row.get("grade"))
            except (TypeError, ValueError):
                continue
            breakdown = {
                str(head): str(value)
                for head, value in (row.get("breakdown") or {}).items()
                if str(value).strip() not in ("", "0")
            }
            structure = FeeStructure(
                school=school, grade=grade, term=term, year=year,
                breakdown=breakdown, amount=Decimal("0"),
            )
            total = structure.total_from_breakdown()
            if total <= 0:
                # An emptied row means "this class is not charged this term".
                # Only drop it if nothing has been billed from it yet.
                gone, _ = FeeStructure.objects.filter(
                    school=school, grade=grade, term=term, year=year,
                    invoice__isnull=True,
                ).delete()
                cleared += bool(gone)
                continue
            FeeStructure.objects.update_or_create(
                school=school, grade=grade, term=term, year=year,
                defaults={
                    "breakdown": breakdown,
                    "amount": total,
                    "description": row.get("description", ""),
                },
            )
            saved += 1
        return Response({"saved": saved, "cleared": cleared})


class GenerateInvoicesView(APIView):
    """POST /api/payments/generate-invoices/ — raise the term's invoices.

    Either {fee_structure} for one class, or {term, year} to bill every class
    that has a fee set — a school sets its whole fee note at once, so it
    should be able to invoice the whole school at once.

    Idempotent: a learner who already has an invoice for a fee structure is
    skipped, so re-running after admitting a new learner bills only them."""

    def post(self, request):
        user = request.user
        if not (user.is_superuser or user.role == "ADMIN"):
            raise PermissionDenied("Only the school office raises fee invoices.")

        from apps.students.models import Learner

        structures = []
        if request.data.get("fee_structure"):
            structures = [get_object_or_404(
                FeeStructure.objects.filter(school=user.school),
                pk=request.data["fee_structure"],
            )]
        else:
            try:
                term = int(request.data.get("term"))
                year = int(request.data.get("year"))
            except (TypeError, ValueError):
                raise ValidationError(
                    {"detail": "Give a fee_structure, or a term and year."}
                )
            structures = list(
                FeeStructure.objects.filter(school=user.school, term=term, year=year)
            )
            if not structures:
                raise ValidationError(
                    {"detail": f"No fee is set for Term {term} {year} yet."}
                )

        stream = (request.data.get("stream") or "").strip()
        created_total = skipped_total = 0
        classes = []
        for structure in structures:
            already = set(
                Invoice.objects.filter(
                    school=user.school, fee_structure=structure
                ).values_list("learner_id", flat=True)
            )
            learners = Learner.objects.filter(
                school=user.school, grade=structure.grade, active=True
            )
            if stream:
                learners = learners.filter(stream=stream)

            created = [
                Invoice(
                    school=user.school,
                    learner=learner,
                    fee_structure=structure,
                    amount_due=structure.amount,
                )
                for learner in learners
                if learner.id not in already
            ]
            Invoice.objects.bulk_create(created)
            created_total += len(created)
            skipped_total += len(already)
            if created:
                classes.append(str(structure))

        return Response(
            {
                "created": created_total,
                "skipped_existing": skipped_total,
                "classes": classes,
                "fee_structure": str(structures[0]) if len(structures) == 1 else "",
            },
            status=status.HTTP_201_CREATED,
        )


class FeeRegisterXlsxView(APIView):
    """GET /api/payments/register.xlsx?... — the filtered fee register as a
    real workbook, one row per learner with what they owe and have paid."""

    def get(self, request):
        user = request.user
        if not (user.is_superuser or user.role in ("ADMIN", "TEACHER")):
            raise PermissionDenied("The fee register is staff-only.")

        rows = (
            Invoice.objects.filter(school=user.school)
            .select_related("learner", "fee_structure")
            .order_by("learner__grade", "learner__stream", "learner__admission_number")
        )
        for param, field in (
            ("learner__grade", "learner__grade"),
            ("learner__stream", "learner__stream"),
            ("fee_structure__term", "fee_structure__term"),
            ("fee_structure__year", "fee_structure__year"),
            ("status", "status"),
        ):
            value = request.query_params.get(param)
            if value not in (None, ""):
                rows = rows.filter(**{field: value})

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Fee register"
        ws.append([
            "Adm No", "Learner", "Grade", "Stream", "Term", "Year",
            "Fee due", "Paid", "Balance", "Status", "Last payment",
        ])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2B6CB0")
        for inv in rows:
            last = inv.payments.order_by("-paid_on").first()
            ws.append([
                inv.learner.admission_number,
                inv.learner.full_name,
                inv.learner.grade,
                inv.learner.stream,
                inv.fee_structure.term,
                inv.fee_structure.year,
                float(inv.amount_due),
                float(inv.amount_paid),
                float(inv.balance),
                inv.get_status_display(),
                last.paid_on.isoformat() if last else "",
            ])
        ws.freeze_panes = "C2"
        for i, width in enumerate([12, 26, 8, 10, 7, 8, 12, 12, 12, 14, 14], start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = 'attachment; filename="fee_register.xlsx"'
        return response


class StkPushView(APIView):
    """Trigger an M-Pesa STK Push for an invoice (defaults to the balance)."""

    def post(self, request):
        serializer = StkPushSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        invoice = get_object_or_404(
            Invoice.objects.filter(school=request.user.school), pk=data["invoice"]
        )
        amount = data.get("amount") or invoice.balance
        if amount <= 0:
            return Response({"detail": "Invoice is already settled."}, status=400)

        try:
            result = daraja.stk_push(
                phone=data["phone"],
                amount=int(amount),
                account_reference=invoice.learner.admission_number,
            )
        except daraja.DarajaError as exc:
            logger.error("STK push failed: %s", exc)
            return Response({"detail": str(exc)}, status=status.HTTP_502_BAD_GATEWAY)

        StkPushRequest.objects.create(
            school=request.user.school,
            invoice=invoice,
            phone=data["phone"],
            amount=amount,
            checkout_request_id=result.get("CheckoutRequestID"),
            merchant_request_id=result.get("MerchantRequestID", ""),
        )
        return Response(result, status=status.HTTP_202_ACCEPTED)


class StkCallbackView(APIView):
    """Daraja STK result webhook. Unauthenticated by necessity; validated by
    matching CheckoutRequestID to a request we actually made, and idempotent
    on the M-Pesa receipt number."""

    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request):
        callback = request.data.get("Body", {}).get("stkCallback", {})
        checkout_id = callback.get("CheckoutRequestID")
        stk_request = StkPushRequest.objects.filter(checkout_request_id=checkout_id).first()
        if not stk_request:
            logger.warning("STK callback for unknown CheckoutRequestID %s", checkout_id)
            return Response({"ResultCode": 0, "ResultDesc": "Accepted"})

        result_code = callback.get("ResultCode")
        stk_request.succeeded = result_code == 0
        stk_request.result_description = callback.get("ResultDesc", "")
        stk_request.save()

        if result_code == 0:
            items = {
                item.get("Name"): item.get("Value")
                for item in callback.get("CallbackMetadata", {}).get("Item", [])
            }
            record_transaction(
                school=stk_request.school,
                source="STK",
                receipt=items.get("MpesaReceiptNumber", f"UNKNOWN-{checkout_id}"),
                phone=str(items.get("PhoneNumber", stk_request.phone)),
                amount=items.get("Amount", stk_request.amount),
                invoice=stk_request.invoice,
                raw=callback,
            )
        return Response({"ResultCode": 0, "ResultDesc": "Accepted"})


class C2BConfirmationView(APIView):
    """Daraja C2B confirmation webhook (parent paid the paybill directly,
    account reference = admission number). Idempotent on TransID."""

    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request):
        data = request.data
        shortcode = str(data.get("BusinessShortCode", ""))
        school = School.objects.first() if not shortcode else (
            School.objects.filter(code=shortcode).first() or School.objects.first()
        )
        if school is None:
            return Response({"ResultCode": 0, "ResultDesc": "Accepted"})

        record_transaction(
            school=school,
            source="C2B",
            receipt=data.get("TransID", ""),
            phone=str(data.get("MSISDN", "")),
            amount=Decimal(str(data.get("TransAmount", "0"))),
            account_reference=data.get("BillRefNumber", ""),
            raw=dict(data),
        )
        return Response({"ResultCode": 0, "ResultDesc": "Accepted"})
