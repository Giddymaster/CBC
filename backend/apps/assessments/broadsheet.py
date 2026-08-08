"""The class broadsheet: every learner × every learning area, in one grid.

The marks book a Kenyan staff room actually works from — one row per learner,
one column per subject (mean percentage and level for the term), then total,
mean and rank. Served as JSON for the screen, as a real .xlsx for Excel, and
the same class (or the whole school) as a single printable PDF of report forms.
"""

from io import BytesIO

from django.http import HttpResponse
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.students.models import Learner

from .models import Score, derive_competency_level
from .pdf import render_report_cards_pdf
from .reports import build_report_card


def _require_staff(user):
    if not (user.is_superuser or user.role in ("ADMIN", "TEACHER")):
        raise PermissionDenied("The broadsheet is staff-only.")


def _class_params(request, grade_required=True):
    grade = request.query_params.get("grade")
    if grade in (None, ""):
        if grade_required:
            raise ValidationError({"grade": "Choose a grade."})
        grade = None
    else:
        grade = int(grade)
    stream = request.query_params.get("stream") or ""
    try:
        term = int(request.query_params.get("term"))
        year = int(request.query_params.get("year"))
    except (TypeError, ValueError):
        raise ValidationError({"detail": "term and year are required numbers."})
    return grade, stream, term, year


def build_broadsheet(school, grade, stream, term, year):
    learners = Learner.objects.filter(school=school, grade=grade, active=True)
    if stream:
        learners = learners.filter(stream=stream)
    learners = list(learners.order_by("admission_number"))

    scores = (
        Score.objects.filter(
            school=school,
            learner__in=[l.id for l in learners],
            assessment__grade=grade,
            assessment__term=term,
            assessment__year=year,
        )
        .select_related("assessment__learning_area")
    )

    # learner -> area -> [percent, ...]
    per_learner = {l.id: {} for l in learners}
    area_names = set()
    for s in scores:
        area = s.assessment.learning_area.name
        area_names.add(area)
        if s.assessment.max_marks:
            per_learner[s.learner_id].setdefault(area, []).append(
                float(s.marks) / s.assessment.max_marks * 100
            )
    areas = sorted(area_names)

    rows = []
    for l in learners:
        cells = {}
        percents = []
        for area in areas:
            marks = per_learner[l.id].get(area)
            if marks:
                pct = sum(marks) / len(marks)
                percents.append(pct)
                cells[area] = {
                    "percent": round(pct, 1),
                    "level": derive_competency_level(pct),
                }
        mean = sum(percents) / len(percents) if percents else 0
        rows.append({
            "learner": l.id,
            "admission_number": l.admission_number,
            "name": l.full_name,
            "stream": l.stream,
            "areas": cells,
            "total": round(sum(percents), 1),
            "mean": round(mean, 1),
        })

    # Rank on the mean, ties sharing a position (1, 2, 2, 4...).
    ordered = sorted(rows, key=lambda r: -r["mean"])
    last_mean, last_rank = None, 0
    for i, row in enumerate(ordered, start=1):
        if row["mean"] != last_mean:
            last_rank, last_mean = i, row["mean"]
        row["rank"] = last_rank if row["mean"] else None

    return {"grade": grade, "stream": stream, "term": term, "year": year,
            "areas": areas, "rows": rows}


class BroadsheetView(APIView):
    """GET /api/report-cards/broadsheet/?grade=&stream=&term=&year="""

    def get(self, request):
        _require_staff(request.user)
        grade, stream, term, year = _class_params(request)
        return Response(
            build_broadsheet(request.user.school, grade, stream, term, year)
        )


class BroadsheetXlsxView(APIView):
    """GET /api/report-cards/broadsheet.xlsx?grade=&stream=&term=&year= —
    the same grid as a real Excel workbook."""

    def get(self, request):
        _require_staff(request.user)
        grade, stream, term, year = _class_params(request)
        data = build_broadsheet(request.user.school, grade, stream, term, year)

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = f"G{grade}{(' ' + stream) if stream else ''} T{term} {year}"[:31]

        head_font = Font(bold=True, color="FFFFFF")
        head_fill = PatternFill("solid", fgColor="2B6CB0")
        headers = (
            ["Adm No", "Name", "Stream"]
            + [f"{a} %" for a in data["areas"]]
            + ["Total", "Mean %", "Rank"]
        )
        ws.append(headers)
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center")
        for row in data["rows"]:
            ws.append(
                [row["admission_number"], row["name"], row["stream"]]
                + [row["areas"].get(a, {}).get("percent") for a in data["areas"]]
                + [row["total"], row["mean"], row["rank"]]
            )
        ws.freeze_panes = "D2"
        widths = [10, 24, 8] + [14] * len(data["areas"]) + [8, 8, 6]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f'attachment; filename="broadsheet_G{grade}'
            f'{("_" + stream) if stream else ""}_T{term}_{year}.xlsx"'
        )
        return response


class ClassReportPdfView(APIView):
    """GET /api/report-cards/class.pdf?grade=&stream=&term=&year= — every
    learner's report form in one printable PDF. Omit grade for the whole
    school (one page per learner, in class order)."""

    def get(self, request):
        _require_staff(request.user)
        grade, stream, term, year = _class_params(request, grade_required=False)
        learners = Learner.objects.filter(
            school=request.user.school, active=True
        ).select_related("pathway", "school")
        if grade is not None:
            learners = learners.filter(grade=grade)
            if stream:
                learners = learners.filter(stream=stream)
        learners = learners.order_by("grade", "stream", "admission_number")

        reports = [
            build_report_card(l, term=term, year=year) for l in learners
        ]
        if not reports:
            raise ValidationError({"detail": "No learners in that class."})
        pdf_bytes = render_report_cards_pdf(reports)
        label = (
            f"G{grade}{('_' + stream) if stream else ''}"
            if grade is not None else "school"
        )
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'inline; filename="report_forms_{label}_T{term}_{year}.pdf"'
        )
        return response
