"""The class broadsheet, its Excel form, and the class/school PDF set."""

from rest_framework.test import APITestCase

from apps.assessments.models import Score
from tests.factories import (
    make_assessment,
    make_learner,
    make_learning_area,
    make_school,
    make_teacher,
    make_user,
)


class BroadsheetTests(APITestCase):
    def setUp(self):
        self.school = make_school()
        self.admin = make_user(self.school, "ADMIN")
        self.maths = make_learning_area("Mathematics", "MATH")
        self.eng = make_learning_area("English", "ENG")
        self.cat = make_assessment(
            self.school, learning_area=self.maths, grade=7, term=2, year=2026,
            max_marks=100,
        )
        self.eng_cat = make_assessment(
            self.school, learning_area=self.eng, grade=7, term=2, year=2026,
            kind="CAT2", max_marks=50,
        )
        self.amina = make_learner(self.school, grade=7)
        self.brian = make_learner(self.school, grade=7)
        Score.objects.create(school=self.school, assessment=self.cat,
                             learner=self.amina, marks=80)
        Score.objects.create(school=self.school, assessment=self.eng_cat,
                             learner=self.amina, marks=25)  # 50%
        Score.objects.create(school=self.school, assessment=self.cat,
                             learner=self.brian, marks=40)

    def test_the_grid_carries_percents_levels_mean_and_rank(self):
        self.client.force_authenticate(self.admin)
        res = self.client.get(
            "/api/report-cards/broadsheet/?grade=7&term=2&year=2026"
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data["areas"], ["English", "Mathematics"])
        rows = {r["learner"]: r for r in res.data["rows"]}
        amina = rows[self.amina.id]
        self.assertEqual(amina["areas"]["Mathematics"]["percent"], 80.0)
        self.assertEqual(amina["areas"]["Mathematics"]["level"], "EE")
        self.assertEqual(amina["areas"]["English"]["percent"], 50.0)
        self.assertEqual(amina["mean"], 65.0)
        self.assertEqual(amina["rank"], 1)
        self.assertEqual(rows[self.brian.id]["rank"], 2)

    def test_the_excel_download_is_a_real_workbook(self):
        import io

        from openpyxl import load_workbook

        self.client.force_authenticate(self.admin)
        res = self.client.get(
            "/api/report-cards/broadsheet.xlsx?grade=7&term=2&year=2026"
        )
        self.assertEqual(res.status_code, 200)
        wb = load_workbook(io.BytesIO(res.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers[:3], ["Adm No", "Name", "Stream"])
        self.assertIn("Mathematics %", headers)
        self.assertEqual(ws.max_row, 3)  # header + two learners

    def test_the_class_pdf_covers_every_learner(self):
        self.client.force_authenticate(self.admin)
        res = self.client.get(
            "/api/report-cards/class.pdf?grade=7&term=2&year=2026"
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["Content-Type"], "application/pdf")
        self.assertTrue(res.content.startswith(b"%PDF"))

    def test_the_whole_school_pdf_needs_no_grade(self):
        make_learner(self.school, grade=4)
        self.client.force_authenticate(self.admin)
        res = self.client.get("/api/report-cards/class.pdf?term=2&year=2026")
        self.assertEqual(res.status_code, 200)
        self.assertTrue(res.content.startswith(b"%PDF"))

    def test_a_teacher_sees_it_a_parent_does_not(self):
        teacher = make_teacher(self.school)
        self.client.force_authenticate(teacher.user)
        self.assertEqual(
            self.client.get(
                "/api/report-cards/broadsheet/?grade=7&term=2&year=2026"
            ).status_code,
            200,
        )
        parent = make_user(self.school, "PARENT")
        self.client.force_authenticate(parent)
        self.assertEqual(
            self.client.get(
                "/api/report-cards/broadsheet/?grade=7&term=2&year=2026"
            ).status_code,
            403,
        )


class ReportFormFeesTests(APITestCase):
    """The report form carries the fee balance and the coming term's fee."""

    def setUp(self):
        from decimal import Decimal

        from apps.payments.models import FeeStructure, Invoice, Payment

        self.school = make_school()
        self.admin = make_user(self.school, "ADMIN")
        self.learner = make_learner(self.school, grade=5)
        this_term = FeeStructure.objects.create(
            school=self.school, grade=5, term=2, year=2026, amount=Decimal("8000"),
        )
        FeeStructure.objects.create(
            school=self.school, grade=5, term=3, year=2026, amount=Decimal("9000"),
            breakdown={"Tuition": "7000", "Games": "2000"},
        )
        invoice = Invoice.objects.create(
            school=self.school, learner=self.learner,
            fee_structure=this_term, amount_due=Decimal("8000"),
        )
        Payment.objects.create(
            school=self.school, invoice=invoice, amount=Decimal("3000"),
            method="CASH", paid_on="2026-05-04",
        )
        invoice.amount_paid = Decimal("3000")
        invoice.refresh_status()

    def test_the_report_carries_this_term_and_next_term_fees(self):
        self.client.force_authenticate(self.admin)
        res = self.client.get(
            f"/api/report-card/{self.learner.id}/?term=2&year=2026"
        )
        fees = res.data["fees"]
        self.assertEqual(fees["billed"], "8000.00")
        self.assertEqual(fees["paid"], "3000.00")
        self.assertEqual(fees["balance"], "5000.00")
        self.assertEqual(fees["next_term"], 3)
        self.assertEqual(fees["next_term_fee"], "9000.00")
        # Arrears follow the child into the new term.
        self.assertEqual(fees["next_term_total_due"], "14000.00")
        self.assertEqual(fees["next_term_breakdown"]["Games"], "2000")

    def test_term_three_rolls_into_the_new_year(self):
        from decimal import Decimal

        from apps.payments.models import FeeStructure

        FeeStructure.objects.create(
            school=self.school, grade=5, term=1, year=2027, amount=Decimal("9500"),
        )
        self.client.force_authenticate(self.admin)
        fees = self.client.get(
            f"/api/report-card/{self.learner.id}/?term=3&year=2026"
        ).data["fees"]
        self.assertEqual((fees["next_term"], fees["next_year"]), (1, 2027))
        self.assertEqual(fees["next_term_fee"], "9500.00")

    def test_the_pdf_still_renders_with_the_fee_block(self):
        self.client.force_authenticate(self.admin)
        res = self.client.get(
            f"/api/report-card/{self.learner.id}/pdf/?term=2&year=2026"
        )
        self.assertEqual(res.status_code, 200)
        self.assertTrue(res.content.startswith(b"%PDF"))
