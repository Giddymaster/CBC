"""The fee register: raising invoices, part-payments, and the export."""

from datetime import date
from decimal import Decimal

from rest_framework.test import APITestCase

from apps.payments.models import FeeStructure, Invoice, Payment
from tests.factories import make_learner, make_school, make_teacher, make_user


class FeeRegisterTests(APITestCase):
    def setUp(self):
        self.school = make_school()
        self.admin = make_user(self.school, "ADMIN")
        self.structure = FeeStructure.objects.create(
            school=self.school, grade=5, term=2, year=2026, amount=Decimal("8000"),
        )
        self.amina = make_learner(self.school, grade=5, stream="North")
        self.brian = make_learner(self.school, grade=5, stream="South")
        make_learner(self.school, grade=7)  # another grade, not billed

    def test_generating_invoices_bills_the_whole_grade_once(self):
        self.client.force_authenticate(self.admin)
        res = self.client.post(
            "/api/payments/generate-invoices/",
            {"fee_structure": self.structure.id}, format="json",
        )
        self.assertEqual(res.status_code, 201, res.data)
        self.assertEqual(res.data["created"], 2)
        self.assertEqual(Invoice.objects.count(), 2)
        self.assertEqual(Invoice.objects.first().amount_due, Decimal("8000"))

        # Re-running after a new admission bills only the newcomer.
        make_learner(self.school, grade=5, stream="North")
        again = self.client.post(
            "/api/payments/generate-invoices/",
            {"fee_structure": self.structure.id}, format="json",
        )
        self.assertEqual(again.data["created"], 1)
        self.assertEqual(Invoice.objects.count(), 3)

    def test_the_register_carries_admission_number_and_class(self):
        self.client.force_authenticate(self.admin)
        self.client.post(
            "/api/payments/generate-invoices/",
            {"fee_structure": self.structure.id}, format="json",
        )
        row = self.client.get("/api/payments/invoices/").data["results"][0]
        self.assertEqual(row["admission_number"], self.amina.admission_number)
        self.assertEqual(row["learner_name"], self.amina.full_name)
        self.assertEqual(row["grade"], 5)
        self.assertEqual(row["term"], 2)
        self.assertEqual(row["year"], 2026)

    def _invoice(self):
        return Invoice.objects.create(
            school=self.school, learner=self.amina,
            fee_structure=self.structure, amount_due=Decimal("8000"),
        )

    def test_instalments_of_different_sizes_add_up(self):
        """A family pays 3,000 cash in May and 5,000 by M-Pesa in June."""
        invoice = self._invoice()
        self.client.force_authenticate(self.admin)
        for amount, method, day in (
            ("3000", "CASH", "2026-05-04"), ("5000", "MPESA", "2026-06-11"),
        ):
            res = self.client.post(
                "/api/payments/payments/",
                {"invoice": invoice.id, "amount": amount, "method": method,
                 "paid_on": day, "reference": "REF1"},
                format="json",
            )
            self.assertEqual(res.status_code, 201, res.data)
        invoice.refresh_from_db()
        self.assertEqual(invoice.amount_paid, Decimal("8000"))
        self.assertEqual(invoice.status, "PAID")
        self.assertEqual(invoice.payments.count(), 2)

    def test_a_part_payment_shows_as_partial_with_the_balance(self):
        invoice = self._invoice()
        self.client.force_authenticate(self.admin)
        self.client.post(
            "/api/payments/payments/",
            {"invoice": invoice.id, "amount": "2500", "method": "BANK",
             "paid_on": "2026-05-20"},
            format="json",
        )
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, "PARTIAL")
        self.assertEqual(invoice.balance, Decimal("5500"))

    def test_reversing_a_payment_re_totals_the_invoice(self):
        invoice = self._invoice()
        self.client.force_authenticate(self.admin)
        created = self.client.post(
            "/api/payments/payments/",
            {"invoice": invoice.id, "amount": "8000", "method": "CASH",
             "paid_on": "2026-05-04"},
            format="json",
        ).data
        self.client.delete(f"/api/payments/payments/{created['id']}/")
        invoice.refresh_from_db()
        self.assertEqual(invoice.amount_paid, Decimal("0"))
        self.assertEqual(invoice.status, "UNPAID")

    def test_a_teacher_cannot_record_a_payment(self):
        invoice = self._invoice()
        teacher = make_teacher(self.school)
        self.client.force_authenticate(teacher.user)
        res = self.client.post(
            "/api/payments/payments/",
            {"invoice": invoice.id, "amount": "1000", "method": "CASH",
             "paid_on": "2026-05-04"},
            format="json",
        )
        self.assertEqual(res.status_code, 403)
        self.assertFalse(Payment.objects.exists())

    def test_the_register_filters_by_class_and_term(self):
        self.client.force_authenticate(self.admin)
        self.client.post(
            "/api/payments/generate-invoices/",
            {"fee_structure": self.structure.id}, format="json",
        )
        res = self.client.get(
            "/api/payments/invoices/?learner__grade=5&learner__stream=North"
            "&fee_structure__term=2"
        )
        self.assertEqual(res.data["count"], 1)
        self.assertEqual(res.data["results"][0]["stream"], "North")

    def test_the_grid_returns_vote_head_columns_for_every_grade(self):
        self.client.force_authenticate(self.admin)
        res = self.client.get("/api/payments/fee-structures/grid/?term=2&year=2026")
        self.assertEqual(res.status_code, 200)
        self.assertIn("Tuition", res.data["columns"])
        self.assertIn("Games", res.data["columns"])
        grades = [r["grade"] for r in res.data["rows"]]
        self.assertIn(5, grades)
        row = next(r for r in res.data["rows"] if r["grade"] == 5)
        self.assertEqual(row["total"], "8000.00")

    def test_saving_the_grid_itemises_and_totals_each_class(self):
        self.client.force_authenticate(self.admin)
        res = self.client.post(
            "/api/payments/fee-structures/grid/",
            {
                "term": 3, "year": 2026,
                "rows": [
                    {"grade": 4, "breakdown": {
                        "Tuition": "6000", "Games": "500", "Lunch": "2500"}},
                    {"grade": 5, "breakdown": {"Tuition": "7000", "Exams": "800"}},
                    {"grade": 6, "breakdown": {}},  # not charged
                ],
            },
            format="json",
        )
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(res.data["saved"], 2)
        g4 = FeeStructure.objects.get(grade=4, term=3, year=2026)
        self.assertEqual(g4.amount, Decimal("9000"))
        self.assertEqual(g4.breakdown["Games"], "500")
        self.assertFalse(
            FeeStructure.objects.filter(grade=6, term=3, year=2026).exists()
        )
        # The saved columns come back on the next read.
        grid = self.client.get(
            "/api/payments/fee-structures/grid/?term=3&year=2026"
        ).data
        self.assertIn("Lunch", grid["columns"])

    def test_a_teacher_cannot_rewrite_the_fee_structure(self):
        teacher = make_teacher(self.school)
        self.client.force_authenticate(teacher.user)
        res = self.client.post(
            "/api/payments/fee-structures/grid/",
            {"term": 3, "year": 2026,
             "rows": [{"grade": 4, "breakdown": {"Tuition": "1"}}]},
            format="json",
        )
        self.assertEqual(res.status_code, 403)

    def test_the_excel_register_downloads(self):
        import io

        from openpyxl import load_workbook

        self.client.force_authenticate(self.admin)
        self.client.post(
            "/api/payments/generate-invoices/",
            {"fee_structure": self.structure.id}, format="json",
        )
        Payment.objects.create(
            school=self.school, invoice=Invoice.objects.first(),
            amount=Decimal("1000"), method="CASH", paid_on=date(2026, 5, 4),
        )
        res = self.client.get("/api/payments/register.xlsx?learner__grade=5")
        self.assertEqual(res.status_code, 200)
        ws = load_workbook(io.BytesIO(res.content)).active
        self.assertEqual(
            [c.value for c in ws[1]][:4], ["Adm No", "Learner", "Grade", "Stream"]
        )
        self.assertEqual(ws.max_row, 3)  # header + two learners
